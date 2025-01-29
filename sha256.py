import hashlib
import openpyxl
import pwinput

wBook = openpyxl.load_workbook("data.xlsx")
sheet = wBook.active

def encrypt(string):
    sha_signature = \
        hashlib.sha256(string.encode()).hexdigest()
    return sha_signature

def login():
    user = input("User: ")
    pswd = pwinput.pwinput()
    encrypted_pswd = encrypt(pswd) 
    user_found = False

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == user:
            user_found = True
            if row[1] == encrypted_pswd:
                print("Welcome Back!")
                return

    if not user_found:
        print("Usuario no registrado!, Intente registrarse...")
    else:
        print("Contraseña incorrecta. Intente de nuevo.")

def register():
    user = input("User: ")
    pswd = pwinput.pwinput()
    encrypted_pswd = encrypt(pswd)

    # Check if the username already exists
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming the first row is the header
        if row[0] == user:  # Check the first column for the username
            print("El nombre de usuario ya existe. Por favor, elige otro.")
            return

    data = [user, encrypted_pswd]
    sheet.append(data)
    wBook.save("data.xlsx")
    print("Registro exitoso.")
    return

print("Quieres iniciar sesión o registrarte? ")
option = int(input("1.Iniciar Sesión \n2.Registrarse\nEliga 1 o 2: "))

if option == 1:
    login()
elif option == 2:
    register()
else:
    print("Opción no reconocida...")
